import time
import random
import openpyxl
from dotenv import load_dotenv
import os
# 載入 selenium 套件
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from openpyxl.utils.exceptions import IllegalCharacterError
from fake_useragent import UserAgent


# 加載 .env 文件的內容到環境變數中
load_dotenv()
# 從環境變數中取得 executable_path 的值
chrome_executable_path  = os.getenv("CHROME_EXECUTABLE_PATH")
print(chrome_executable_path)
# 爬蟲目標網頁
TARGET_URL = 'https://www.104.com.tw/jobs/search/?ro=0&keyword=%E5%89%8D%E7%AB%AF%E5%B7%A5%E7%A8%8B%E5%B8%AB&expansionType=area%2Cspec%2Ccom%2Cjob%2Cwf%2Cwktm&area=6001001000&order=15&asc=0&page=1&jobexp=1&mode=s&jobsource=2018indexpoc&langFlag=0&langStatus=0&recommendJob=1&hotJob=1'

# 設定隨機 User-Agent
ua = UserAgent()
options = Options()
options.add_argument(f"user-agent={ua.random}")
  # 設定 Chrome Driver 的執行檔案路徑
options.chrome_executable_path = chrome_executable_path
# options.add_argument("--headless")
# 建立 Driver 物件實體，用程式操作瀏覽器用作
browser = webdriver.Chrome(options = options)

# 訪問指定104 頁面
browser.get(TARGET_URL)  

# 等待頁面載入並找到排序下拉菜單
wait = WebDriverWait(browser, 10)
dropdown_element = wait.until(EC.presence_of_element_located((By.ID, "js-sort")))


# 選擇下拉框中的「日期排序」
dropdown = Select(dropdown_element)
dropdown.select_by_value("2-0")

# 等待一段隨機時間
time.sleep(random.uniform(2, 5))


# 設定滾動次數限制（視情況調整）
# 初始化變數
scrolls = 0
last_count = 0
MAX_SCROLLS = 50

while scrolls < MAX_SCROLLS:
    # 模擬滾動到頁面底部
    browser.execute_script("window.scrollBy(0, 1500);")  # 這裡改為滾動800像素，模擬真實滾動
    time.sleep(random.uniform(1.5, 3.5))
    print('滾動了', scrolls , '次')
    # 等待新的 `article` 元素出現
    articles = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "article.js-job-item")))
    time.sleep(random.uniform(2, 5))
    last_count = len(articles)
    scrolls += 1

clicks = 1
# 當滾動到第15頁之後，開始手動點擊button
previous_button = None
while True:
    print('開始搜尋more page 按鈕')
    try:
        # 獲取所有的button.js-more-page元素
        load_more_buttons = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "button.js-more-page")))
        # 選取最後一個按鈕
        latest_load_more_button = load_more_buttons[-1]
        # 如果最後一個按鈕與上次的按鈕相同，跳出迴圈
        if latest_load_more_button == previous_button:
            print('沒有新的 more page 按鈕，結束迴圈')
            break
        # 使用ActionChains點擊該按鈕，確保其可見和可點擊
        actions = ActionChains(browser)
        actions.move_to_element(latest_load_more_button).click().perform()
        print('點擊第', clicks, '次')
        time.sleep(5)  # 給予頁面更多的時間加載
        for _ in range(3):
            browser.execute_script("window.scrollBy(0, 1600);")
            time.sleep(random.uniform(1.5, 3.5))
        # 更新 previous_button
        previous_button = latest_load_more_button
        clicks += 1
        time.sleep(2.5)  # 給予頁面更多的時間加載
    except:
        # 當找不到任何符合button.js-more-page元素，跳出循環
        print('找不到more page 按鈕')
        break
# 確保文章元素已經出現
articles = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "article.js-job-item")))

# 用於儲存所有資訊的列表
jobs = []

for article in articles:
    # 獲取所需的資訊
    job_title = article.find_element(By.CSS_SELECTOR, "a.js-job-link").text
    job_link = article.find_element(By.CSS_SELECTOR, "a.js-job-link").get_attribute('href')
    company_name = article.find_element(By.CSS_SELECTOR, "ul.b-list-inline li a").text
    location = article.find_element(By.CSS_SELECTOR, "ul.job-list-intro li:nth-child(1)").text
    experience = article.find_element(By.CSS_SELECTOR, "ul.job-list-intro li:nth-child(3)").text
    # 使用 try-except 來捕獲可能出現的異常
    try:
        job_description = article.find_element(By.CSS_SELECTOR, "p.job-list-item__info").text
    except NoSuchElementException:
        print("Element not found!")
        job_description = "N/A"

    # 使用 try-except 來捕獲可能出現的異常
    try:
        salary = article.find_element(By.CSS_SELECTOR, "div.job-list-tag span.b-tag--default").text
    except:
        salary = 'N/A'
    
    # 儲存到字典並加入列表
    job_info = {
        "title": job_title,
        "link": job_link,
        "company": company_name,
        "location": location,
        "experience": experience,
        "description": job_description,
        "salary": salary
    }
    jobs.append(job_info)


# 先建立一個新的 Excel 工作簿和工作表
print('開始建立excel')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Job Listings"

# 寫入標題列 (headers)
headers = ['title', 'link', 'company', 'location', 'experience', 'description', 'salary']
for col_num, header in enumerate(headers, 1):
    col_letter = openpyxl.utils.get_column_letter(col_num)
    ws[f"{col_letter}1"] = header
    ws[f"{col_letter}1"].font = openpyxl.styles.Font(bold=True)

# 填充數據
for row_num, job in enumerate(jobs, 2):  # 從第二行開始，因為第一行是標題列
    ws.cell(row=row_num, column=1, value=job['title'])
    ws.cell(row=row_num, column=2, value=job['link'])
    ws.cell(row=row_num, column=3, value=job['company'])
    ws.cell(row=row_num, column=4, value=job['location'])
    ws.cell(row=row_num, column=5, value=job['experience'])
    try:
        ws.cell(row=row_num, column=6, value=job['description'])
    except IllegalCharacterError:
        print(f"Illegal character found in job description: {job['description']}. Skipping...")
    ws.cell(row=row_num, column=7, value=job['salary'])

# 儲存到指定的 Excel 文件
wb.save("jobs_list.xlsx")
# 關閉瀏覽器
# browser.close()
print('完成')
input("Press Enter to close the browser...")
